"""
33_nj_enrollment.py — real per-school enrollment for NJ from NJ DOE Fall
Enrollment, plus whole-school demographics, and a 5-year history.

Source: https://www.nj.gov/education/doedata/enr/  (one zip per year; the
"School" sheet is district+school level, header on row 3). Years pulled match
the Florida tool's 5-year window exactly: 2021-22 … 2025-26.

This upgrades the NJ side in two ways over the assessment file alone:

  1. `enrollment_2526` is true total K-12 enrollment, so NJ map bubbles size on
     the same basis as Florida's, and the detail card gets a real 5-year
     enrollment history chart instead of a tested-count proxy.

  2. Race / low-income / multilingual shares are recomputed from this
     whole-school survey rather than from grades 3-8 test-takers, matching how
     Florida sources those (FL DOE Membership survey) — so the two regions'
     demographic bars mean the same thing. NJ's low-income measure is Free +
     Reduced Lunch, which is the closest available analog to Florida's
     "economically disadvantaged" but is not computed identically; the UI
     labels it accordingly.

Students-with-disabilities is NOT in this file, so ese_pct keeps the NJSLA
tested-basis value from etl/32 (still better than Florida, which publishes none).

Reads:  data/raw/nj_enrollment/<year>/*.xlsx  (fetched by hand or by this script)
Writes: data/processed/nj_enrollment.json          keyed by school id
        data/processed/nj_schools.geojson          (enrollment_2526 filled in)
        data/processed/nj_school_performance.json  (demographics rebased)
"""
import glob, json, os, sys, urllib.request, zipfile

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
PROC = os.path.join(ROOT, "data", "processed")
RAWDIR = os.path.join(ROOT, "data", "raw", "nj_enrollment")

YEARS = ["2122", "2223", "2324", "2425", "2526"]
URLS = {
    "2526": "https://www.nj.gov/education/doedata/enr/enr26/Enrollment_2526.zip",
    "2425": "https://www.nj.gov/education/doedata/enr/enr25/enrollment_2425.zip",
    "2324": "https://www.nj.gov/education/doedata/enr/enr24/enrollment_2324.zip",
    "2223": "https://www.nj.gov/education/doedata/enr/enr23/enrollment_2223.zip",
    "2122": "https://www.nj.gov/education/doedata/enr/enr22/enrollment_2122.zip",
}

COL = {  # header label -> our field (header row is row 3)
    "District Code": "dcode", "School Code": "scode", "School Name": "name",
    "Total Enrollment": "total",
    "White": "white", "Black": "black", "Hispanic": "hispanic", "Asian": "asian",
    "Native American": "amind", "Hawaiian Native": "pacific", "Two or More Races": "two_plus",
    "Free Lunch": "free_lunch", "Reduced Lunch": "reduced_lunch",
    "Multilingual Learners": "mll",
}
GRADE_COLS = ["Kindergarten Halfday", "Kindergarten Fullday", "First Grade", "Second Grade",
              "Third Grade", "Fourth Grade", "Fifth Grade", "Sixth Grade", "Seventh Grade",
              "Eighth Grade"]


def ensure_year(year):
    d = os.path.join(RAWDIR, year)
    if glob.glob(os.path.join(d, "*.xlsx")):
        return glob.glob(os.path.join(d, "*.xlsx"))[0]
    os.makedirs(RAWDIR, exist_ok=True)
    zpath = os.path.join(RAWDIR, f"{year}.zip")
    if not os.path.exists(zpath):
        print(f"    downloading {year}…")
        req = urllib.request.Request(URLS[year], headers={"User-Agent": "kipp-demographics-etl/1.0"})
        with urllib.request.urlopen(req, timeout=180) as r, open(zpath, "wb") as f:
            f.write(r.read())
    with zipfile.ZipFile(zpath) as z:
        z.extractall(d)
    hits = glob.glob(os.path.join(d, "*.xlsx"))
    if not hits:
        sys.exit(f"No xlsx inside {zpath}")
    return hits[0]


def num(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except ValueError:
        return None


def parse_year(year):
    path = ensure_year(year)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "School" not in wb.sheetnames:
        sys.exit(f"{path}: no 'School' sheet (sheets={wb.sheetnames})")
    ws = wb["School"]
    rows = ws.iter_rows(values_only=True)
    header = None
    for r in rows:
        if r and r[0] == "County Code":
            header = list(r)
            break
    if not header:
        sys.exit(f"{path}: couldn't find header row in 'School' sheet")
    idx = {}
    for label, field in COL.items():
        if label in header:
            idx[field] = header.index(label)
    gidx = [header.index(g) for g in GRADE_COLS if g in header]

    out = {}
    for r in rows:
        if not r or not r[idx["dcode"]]:
            continue
        dcode = str(r[idx["dcode"]]).strip().zfill(4)
        scode = str(r[idx["scode"]]).strip().zfill(3)
        rec = {"name": r[idx["name"]] if "name" in idx else None,
               "total": num(r[idx["total"]])}
        for f in ["white", "black", "hispanic", "asian", "amind", "pacific", "two_plus",
                  "free_lunch", "reduced_lunch", "mll"]:
            if f in idx:
                rec[f] = num(r[idx[f]])
        k8 = sum((num(r[i]) or 0) for i in gidx)
        rec["k8"] = k8 or None
        out[f"{dcode}|{scode}"] = rec
    return out


def main():
    print("[1/4] Parsing NJ DOE Fall Enrollment, 5 years")
    per_year = {}
    for y in YEARS:
        per_year[y] = parse_year(y)
        print(f"    {y}: {len(per_year[y])} schools statewide")

    print("[2/4] Joining to NJ school points")
    with open(os.path.join(PROC, "nj_schools.geojson")) as f:
        schools = json.load(f)
    latest = per_year["2526"]

    enrollment, matched = {}, 0
    for feat in schools["features"]:
        p = feat["properties"]
        key = f"{str(p['district_code']).zfill(4)}|{str(p['school_num']).zfill(3)}"
        # Names only exist in the NJSLA file (the SPR directory carries no name),
        # so non-tested schools arrive here unnamed — this file has names for them.
        cur = latest.get(key) or {}
        if cur.get("name") and (not p.get("name") or "|" in str(p.get("name"))):
            p["name"] = cur["name"]
        years = {}
        for y in YEARS:
            rec = per_year[y].get(key)
            if rec and rec.get("total"):
                years[y] = {"total": rec["total"]}
        if not years:
            continue
        matched += 1
        totals = [years.get(y, {}).get("total") for y in YEARS]
        present = [t for t in totals if t is not None]
        first, last = present[0], present[-1]
        enrollment[p["id"]] = {
            "years": years,
            "change_5yr_n": (last - first) if len(present) > 1 else None,
            "change_5yr_pct": ((last - first) / first) if len(present) > 1 and first else None,
        }
        p["enrollment_2526"] = cur.get("total")
        p["enrollment_k8"] = cur.get("k8")

    print(f"    matched {matched}/{len(schools['features'])} schools to enrollment")

    # Drop entries that are unidentifiable in every source: no real name, no
    # enrollment, no assessment results. These come from stale geocode rows and
    # would otherwise render as anonymous dots on the map.
    with open(perf_path_for_filter := os.path.join(PROC, "nj_school_performance.json")) as f:
        perf_ids = set(json.load(f).keys())
    before = len(schools["features"])
    kept = []
    for feat in schools["features"]:
        p = feat["properties"]
        has_name = p.get("name") and "|" not in str(p["name"])
        if has_name or p["id"] in perf_ids or p.get("enrollment_2526"):
            kept.append(feat)
    schools["features"] = kept
    dropped = before - len(kept)
    print(f"    dropped {dropped} unidentifiable entries (no name, no enrollment, no NJSLA)")
    print(f"    kept {len(kept)} NJ schools")

    with open(os.path.join(PROC, "nj_enrollment.json"), "w") as f:
        json.dump(enrollment, f)
    with open(os.path.join(PROC, "nj_schools.geojson"), "w") as f:
        json.dump(schools, f)

    print("[3/4] Rebasing performance demographics onto whole-school enrollment")
    perf_path = os.path.join(PROC, "nj_school_performance.json")
    with open(perf_path) as f:
        perf = json.load(f)
    by_id = {p["properties"]["id"]:
             f"{str(p['properties']['district_code']).zfill(4)}|{str(p['properties']['school_num']).zfill(3)}"
             for p in schools["features"]}
    rebased = 0
    for sid, rec in perf.items():
        cur = latest.get(by_id.get(sid, ""))
        if not cur or not cur.get("total"):
            continue
        total = cur["total"]
        race = {f: cur.get(f) for f in
                ["white", "black", "hispanic", "asian", "amind", "pacific", "two_plus"]
                if cur.get(f)}
        rec["enrollment"] = total
        rec["race_total"] = total
        rec["race"] = race or None
        rec["race_pct"] = ({k: round(100.0 * v / total, 1) for k, v in race.items()}
                           if race else None)
        frl = (cur.get("free_lunch") or 0) + (cur.get("reduced_lunch") or 0)
        rec["ed_pct"] = round(100.0 * frl / total, 1) if frl else None
        rec["ed_basis"] = "Free + Reduced Lunch (NJ DOE Fall Enrollment 2025-26)"
        if cur.get("mll") is not None:
            rec["ell_pct"] = round(100.0 * cur["mll"] / total, 1)
        rec["demographics_basis"] = "whole school"
        rebased += 1
    with open(perf_path, "w") as f:
        json.dump(perf, f)
    print(f"    rebased {rebased}/{len(perf)} performance records")

    print("[4/4] Sanity check — largest NJ schools by 2025-26 enrollment")
    top = sorted(((p["properties"].get("enrollment_2526") or 0, p["properties"]["name"],
                   p["properties"]["city"]) for p in schools["features"]), reverse=True)[:6]
    for n, nm, city in top:
        print(f"    {n:>6,.0f}  {city:9} {nm[:46]}")
    tot = {}
    for p in schools["features"]:
        c = p["properties"]["city"]
        tot[c] = tot.get(c, 0) + (p["properties"].get("enrollment_2526") or 0)
    print("    city totals (schools in this tool):",
          ", ".join(f"{k}={v:,.0f}" for k, v in sorted(tot.items())))
    print("\nDone.")


if __name__ == "__main__":
    main()
