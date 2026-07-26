"""
Parse a FL DOE "School Grades" statewide xlsx (e.g. SchoolGrades26.xlsx) into
the same per-school performance schema already used by data/processed/
school_performance.json and orange_school_performance.json, and refresh both
files with the new proficiency scores + letter grades.

Why header-name lookup instead of fixed column positions: FL DOE's template is
stable year to year but not guaranteed identical, and columns for prior-year
grades shift as a new year is appended — matching by header text is robust to
minor reordering and (unlike a fixed index) fails loudly if a column we need
disappears, instead of silently reading the wrong one.

Race/ELL/enrollment fields are NOT touched here — those come from the
membership survey files (2526MembBySchoolByGradeByRace.xlsx /
2526ELLCodeMemb.xlsx), which are already 2025-26 vintage and don't need to
change just because a new School Grades release came out. This script only
refreshes proficiency, letter grades, ed_pct, charter/title1/school_type —
then merges those onto the existing race/ELL/enrollment fields already
sitting in data/processed/{school_performance,orange_school_performance}.json.

Usage:
    python3 etl/24_parse_school_grades.py path/to/SchoolGrades26.xlsx

Districts kept: 06 (Broward), 13 (Miami-Dade) -> school_performance.json
                48 (Orange)                   -> orange_school_performance.json
"""
import json, os, re, sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
PROC = os.path.join(ROOT, "data", "processed")

DISTRICTS = {"06": "school_performance.json", "13": "school_performance.json", "48": "orange_school_performance.json"}

HEADER_MAP = {
    "district": "District Number",
    "school_num": "School Number",
    "name": "School Name",
    "ela": "English Language Arts Achievement",
    "math": "Mathematics Achievement",
    "science": "Science Achievement",
    "social_studies": "Social Studies Achievement",
    "pct_tested": "Percent Tested",
    "charter": "Charter School",
    "title1": "Title I",
    "school_type": "School Type",
    "ed_pct": "Percent of Economically Disadvantaged Students",
}
GRADE_COL_RE = re.compile(r"^(?:Grade|Informational Baseline Grade)\s+(\d{4})$")


def find_header_row(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
        if row and row[0] == "District Number":
            return i + 1, row  # openpyxl rows are 1-indexed
    raise SystemExit("Couldn't find the header row (looking for 'District Number' in column A).")


def mean(vals):
    vals = [v for v in vals if v is not None]
    return round(sum(vals) / len(vals), 1) if vals else None


def yesno(v):
    return str(v).strip().upper() == "YES"


def main():
    if len(sys.argv) != 2:
        sys.exit("Usage: python3 etl/24_parse_school_grades.py path/to/SchoolGradesNN.xlsx")
    path = sys.argv[1]
    if not os.path.exists(path):
        sys.exit(f"Not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row_idx, header = find_header_row(ws)

    col = {}
    for key, label in HEADER_MAP.items():
        if label not in header:
            sys.exit(f"Expected column '{label}' not found in {path}. "
                     f"FL DOE may have renamed it — update HEADER_MAP.")
        col[key] = header.index(label)

    grade_cols = {}  # year (int) -> column index
    for i, h in enumerate(header):
        if not isinstance(h, str):
            continue
        m = GRADE_COL_RE.match(h.strip())
        if m:
            grade_cols[int(m.group(1))] = i
    if not grade_cols:
        sys.exit("No 'Grade YYYY' columns found — check the header row format.")
    years = sorted(grade_cols, reverse=True)[:4]  # current year + 3 prior, matching existing schema depth
    print(f"Grade-year columns found: {sorted(grade_cols)}  (keeping {years})")

    by_district = {"06": {}, "13": {}, "48": {}}
    n_rows = 0
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or not row[col["district"]]:
            continue
        dist = str(row[col["district"]]).strip().zfill(2)
        if dist not in DISTRICTS:
            continue
        snum = str(row[col["school_num"]] or "").strip().zfill(4)
        if not snum:
            continue
        n_rows += 1

        def g(key):
            v = row[col[key]]
            return v if v not in ("", None) else None

        grades = {f"g{y}": (row[grade_cols[y]] or None) for y in years}

        by_district[dist][f"{dist}-{snum}"] = {
            "name": g("name"),
            "charter": yesno(row[col["charter"]]),
            "title1": yesno(row[col["title1"]]),
            "school_type": g("school_type"),
            "ela": g("ela"), "math": g("math"), "science": g("science"), "social_studies": g("social_studies"),
            "ela_math": mean([g("ela"), g("math")]),
            "pct_tested": g("pct_tested"),
            "grades": grades,
            "ed_pct": g("ed_pct"),
        }

    print(f"Parsed {n_rows} rows for districts {sorted(DISTRICTS)}: "
          f"{[(d, len(by_district[d])) for d in DISTRICTS]}")

    # Merge onto the existing processed files, keeping their race/ELL/enrollment
    # fields (already 2025-26 vintage) and only overwriting the fields above.
    by_outfile = {}
    for dist, fresh in by_district.items():
        outfile = DISTRICTS[dist]
        by_outfile.setdefault(outfile, {}).update(fresh)

    latest_year = years[0]
    for outfile, fresh in by_outfile.items():
        path_out = os.path.join(PROC, outfile)
        existing = {}
        if os.path.exists(path_out):
            with open(path_out) as f:
                existing = json.load(f)

        updated = 0
        for key, rec in fresh.items():
            old = existing.get(key, {})
            grade_fields = {f"grade_{y}": rec["grades"].get(f"g{y}") for y in years}
            merged = {
                **old,  # keeps race/race_pct/race_total/enrollment/ell_count/ell_pct/ese_pct
                "name": rec["name"] or old.get("name"),
                "charter": rec["charter"],
                "title1": rec["title1"],
                "school_type": rec["school_type"],
                "ela": rec["ela"], "math": rec["math"],
                "science": rec["science"], "social_studies": rec["social_studies"],
                "ela_math": rec["ela_math"],
                "pct_tested": rec["pct_tested"],
                "ed_pct": rec["ed_pct"],
                # Vintage of the proficiency/grade fields above. Stamped per-record
                # rather than assumed globally by the UI, because a school present in
                # last year's release can be absent from this one (closed, or no grade
                # issued) — those keep older numbers, and the UI must not caption them
                # as current.
                "data_year": latest_year,
                **grade_fields,
            }
            existing[key] = merged
            updated += 1

        # Backfill data_year on records this release didn't refresh, inferring it
        # from the most recent grade_YYYY they actually carry.
        stale = 0
        for key, rec in existing.items():
            if rec.get("data_year"):
                continue
            grade_years = [int(m.group(1)) for m in
                           (re.match(r"^grade_(\d{4})$", k) for k in rec) if m]
            have = [y for y in sorted(grade_years, reverse=True) if rec.get(f"grade_{y}")]
            rec["data_year"] = have[0] if have else None
            stale += 1

        with open(path_out, "w") as f:
            json.dump(existing, f)
        print(f"-> {path_out}: {updated} schools refreshed with {latest_year} data "
              f"({len(existing)} total in file)")
        if stale:
            print(f"   {stale} not in this release — kept prior-year data, "
                  f"data_year stamped accordingly")

    print(f"\nDone. Latest school year in this release: {latest_year}.")
    print("Remember to update the in-app 'FL DOE 20XX-YY' captions in web/campus.js, "
          "web/heatmap.js, web/app.js and the grade-trend display in campus.js "
          "(PerformanceBlock's `grades` array + the big letter-grade tile).")


if __name__ == "__main__":
    main()
